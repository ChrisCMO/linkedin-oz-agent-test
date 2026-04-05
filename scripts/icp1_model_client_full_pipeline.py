#!/usr/bin/env python3
"""
Full ICP pipeline for 4 VWC model clients → Excel output.

Steps:
  1. Google Places verification
  2. Apollo org enrichment (company data)
  3. ZoomInfo contact search (free)
  4. Apollo contact search (free) + cross-match ZoomInfo→Apollo
  5. Apollo person enrichment (1 credit each)
  6. LinkedIn validation (Apify: profile scraper + posts scraper + company page)
  7. AI scoring (GPT-5.4)
  8. Connection notes (Adrienne + Melinda) + 3-message sequences
  9. Excel output with prospect sheet + ZoomInfo vs Apollo comparison sheet
"""

import sys, os, time, json, csv, re, random
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from dotenv import load_dotenv
load_dotenv()

import requests
import logging
from datetime import datetime, timedelta
from lib.apollo import ApolloClient
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger()

# ── Clients ──
apollo = ApolloClient()
APIFY_TOKEN = os.environ["APIFY_API_KEY"]
OPENAI_API_KEY = os.environ["OPENAI_API_KEY"]

# Correct Apify actor IDs (skill spec)
PROFILE_SCRAPER = "LpVuK3Zozwuipa5bp"   # Live profile — role verification
POSTS_SCRAPER   = "RE0MriXnFhR3IgVnJ"   # Posts/reposts/feed activity
COMPANY_SCRAPER = "UwSdACBp7ymaGUJjS"   # Company LinkedIn page

GP_HEADERS = {
    "X-Goog-Api-Key": os.environ["GOOGLE_API_KEY"],
    "X-Goog-FieldMask": "places.displayName,places.formattedAddress,places.nationalPhoneNumber,places.websiteUri,places.rating,places.userRatingCount,places.businessStatus",
    "Content-Type": "application/json",
}

# ZoomInfo auth
log.info("Authenticating ZoomInfo...")
zi_resp = requests.post("https://api.zoominfo.com/authenticate", json={
    "username": os.environ["ZOOMINFO_USERNAME"],
    "password": os.environ["ZOOMINFO_PASSWORD"],
})
zi_jwt = zi_resp.json()["jwt"]
zi_headers = {"Authorization": f"Bearer {zi_jwt}", "Content-Type": "application/json"}
log.info("ZoomInfo authenticated.")


# ── Model Clients ──
MODEL_CLIENTS = [
    {
        "name": "Formost Fuji Corporation",
        "domain": "formostfuji.com",
        "industry": "Manufacturing",
        "icp": "Audit & Tax",
        "notes": "Long-tenured client. ~100 employees. Ownership/leadership transition.",
        "zi_search": "Formost Fuji",
        "gp_search": "Formost Fuji Corporation",
    },
    {
        "name": "Shannon & Wilson",
        "domain": "shannonwilson.com",
        "industry": "Professional services (engineering)",
        "icp": "Audit & Tax + Benefit Plan",
        "notes": "ESOP. Top 5 VWC client. 2 benefit plan audits + tax + review.",
        "zi_search": "Shannon Wilson",
        "gp_search": "Shannon Wilson engineering Seattle",
    },
    {
        "name": "Skills Inc.",
        "domain": "skillsinc.com",
        "industry": "Nonprofit / Aerospace manufacturing",
        "icp": "Audit & Tax + Benefit Plan",
        "notes": "Nonprofit like for-profit. Boeing airplane parts.",
        "zi_search": "Skills Inc",
        "gp_search": "Skills Inc Auburn Washington",
    },
    {
        "name": "Carillon Properties",
        "domain": "carillonpoint.com",
        "industry": "Commercial real estate / Hospitality",
        "icp": "Audit & Tax",
        "notes": "CRE + hotel. Old family money. Previously Deloitte.",
        "zi_search": "Carillon Properties",
        "gp_search": "Carillon Properties Kirkland Washington",
    },
]

FINANCE_TITLES_ZI = "CFO OR Chief Financial Officer OR Controller OR VP Finance OR Director of Finance OR Treasurer OR President OR Owner"
FINANCE_TITLES_APOLLO = [
    "CFO", "Chief Financial Officer", "Controller", "VP Finance",
    "Director of Finance", "President", "Owner", "CEO", "Executive Director", "Treasurer",
]

ICP_CONFIG = {
    "target_titles": ["CFO", "Chief Financial Officer", "Controller", "VP Finance", "Director of Finance", "Owner", "President", "CEO"],
    "target_seniorities": ["c_suite", "vp", "director", "owner"],
    "target_industries": [
        "manufacturing", "machinery", "mechanical or industrial engineering",
        "building materials", "aviation & aerospace", "automotive",
        "construction", "civil engineering", "real estate", "property management",
        "professional services", "hospitality", "hotels",
        "nonprofit", "nonprofit organization management",
    ],
    "target_locations": ["Seattle", "Washington", "Pacific Northwest", "Oregon"],
    "employee_count_ranges": ["11-50", "51-200", "201-500", "501-1000"],
    "scoring_config": {
        "revenue_ranges": ["$5M-$150M"],
        "hard_exclusions": ["revenue > $150M", "government", "banking", "Fortune 500"],
        "custom_notes": "These are VWC's existing model clients. Score generously — they ARE the ICP.",
    },
}


# ── Apify Helper ──
def run_actor(actor_id, payload, max_wait=180):
    """Run an Apify actor and return results."""
    headers = {"Authorization": f"Bearer {APIFY_TOKEN}", "Content-Type": "application/json"}
    try:
        r = requests.post(f"https://api.apify.com/v2/acts/{actor_id}/runs", headers=headers, json=payload, timeout=30)
        if r.status_code != 201:
            log.warning(f"  Apify start failed: {r.status_code} {r.text[:200]}")
            return []
        run_id = r.json()["data"]["id"]
        dataset_id = r.json()["data"]["defaultDatasetId"]
        for _ in range(max_wait // 5):
            time.sleep(5)
            sr = requests.get(f"https://api.apify.com/v2/actor-runs/{run_id}", headers=headers, timeout=15)
            status = sr.json()["data"]["status"]
            if status in ("SUCCEEDED", "FAILED", "ABORTED"):
                break
        items = requests.get(f"https://api.apify.com/v2/datasets/{dataset_id}/items", headers=headers, timeout=15).json()
        return items if isinstance(items, list) else []
    except Exception as e:
        log.warning(f"  Apify error: {e}")
        return []


def parse_date(ds):
    if not ds:
        return None
    try:
        return datetime.fromisoformat(ds.replace("Z", "+00:00")).replace(tzinfo=None)
    except (ValueError, TypeError):
        return None


def openai_call(system_prompt, user_content, model="gpt-5.4", temperature=0.8, max_tokens=1000):
    """Direct OpenAI API call."""
    resp = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"},
        json={
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content},
            ],
            "temperature": temperature,
            "max_completion_tokens": max_tokens,
        },
        timeout=60,
    )
    return resp.json()["choices"][0]["message"]["content"].strip()


# ── Connection Note Prompt ──
CONNECTION_NOTE_PROMPT = """You are {sender_name}, an audit partner at VWC CPAs in Seattle. Write a short LinkedIn connection request note TO the prospect below.

Rules:
- MUST be under 200 characters
- Address by first name
- Personalize to role, company, or industry
- Warm, professional, not salesy
- Mention a reason to connect (shared industry interest, their company, their role)
- No emojis, no clichés
- Do NOT mention VWC CPAs by name
- Sound human, not automated

Positioning context (use naturally, don't force):
- VWC has been in business 50 years — partner-level attention, same partners year after year
- Boutique care with regional/national firm expertise
- Never name competitors

Return ONLY the note text."""


# ── Message Sequence Prompt ──
MESSAGE_PROMPT = """You are writing LinkedIn follow-up messages for {sender_name}, an audit partner at VWC CPAs in Seattle.

Generate exactly 3 messages for a drip sequence to the prospect below.

Message 1 (after connection accepted): Reference something specific about them. Short, warm. 2-4 sentences.
Message 2 (~2 weeks later, no reply): Different angle — industry insight or pain point. Soft mention of how VWC helps. 2-4 sentences.
Message 3 (~4 weeks): Final light touch. Brief question. 1-2 sentences.

Positioning (weave naturally):
- VWC: 50 years in business, partner-level attention, same partners year after year
- Boutique care, regional/national firm expertise
- Counter-position to big firms (clients feel under-served) but NEVER name competitors
- Research-driven personalization required

Rules: No emojis. No clichés. Under 300 chars each. Sound human.

Return JSON array: [{{"step": 1, "text": "..."}}, {{"step": 2, "text": "..."}}, {{"step": 3, "text": "..."}}]
Return ONLY valid JSON."""


# ══════════════════════════════════════════════════════════════
#  MAIN PIPELINE
# ══════════════════════════════════════════════════════════════

all_prospects = []       # Final prospect rows
zi_vs_apollo = []        # Comparison data
company_data_map = {}    # company name → company-level data
no_prospect_companies = []

for mc in MODEL_CLIENTS:
    log.info(f"\n{'=' * 70}")
    log.info(f"  COMPANY: {mc['name']}")
    log.info(f"{'=' * 70}")

    co = {
        "name": mc["name"],
        "industry": mc["industry"],
        "icp": mc["icp"],
        "notes": mc["notes"],
        "domain": mc["domain"],
    }

    # ── Step 1: Google Places ──
    log.info("  [1/9] Google Places...")
    gp_resp = requests.post(
        "https://places.googleapis.com/v1/places:searchText",
        headers=GP_HEADERS,
        json={"textQuery": mc["gp_search"]},
    )
    gp_places = gp_resp.json().get("places", [])
    if gp_places:
        gp = gp_places[0]
        co["gp_address"] = gp.get("formattedAddress", "")
        co["gp_phone"] = gp.get("nationalPhoneNumber", "")
        co["gp_website"] = gp.get("websiteUri", "")
        co["gp_rating"] = gp.get("rating", "")
        co["gp_reviews"] = gp.get("userRatingCount", "")
        co["gp_status"] = gp.get("businessStatus", "")
        log.info(f"    Found: {gp.get('displayName', {}).get('text')} | {co['gp_address']}")
    else:
        co.update({"gp_address": "", "gp_phone": "", "gp_website": "", "gp_rating": "", "gp_reviews": "", "gp_status": ""})
        log.info("    Not found in Google Places")

    # ── Step 2: Apollo Org Enrichment ──
    log.info("  [2/9] Apollo Org Enrichment...")
    org_result = apollo._request("POST", "/api/v1/organizations/enrich", json_body={"domain": mc["domain"]})
    org = org_result.get("organization", {})
    co["apollo_company_name"] = org.get("name", "")
    co["apollo_industry"] = org.get("industry", "")
    co["apollo_employees"] = org.get("estimated_num_employees", "")
    rev = org.get("annual_revenue")
    co["apollo_revenue"] = f"${rev / 1e6:.0f}M" if rev and rev >= 1e6 else (str(rev) if rev else "")
    co["apollo_city"] = f"{org.get('city', '')}, {org.get('state', '')}".strip(", ")
    co["company_linkedin"] = org.get("linkedin_url", "")
    co["company_website"] = org.get("website_url", "")
    co["apollo_founded"] = org.get("founded_year", "")
    log.info(f"    {co['apollo_company_name']} | {co['apollo_industry']} | emp={co['apollo_employees']} | rev={co['apollo_revenue']}")
    time.sleep(1)

    # ── Step 3: ZoomInfo Contact Search ──
    log.info("  [3/9] ZoomInfo Contacts...")
    zi_resp2 = requests.post("https://api.zoominfo.com/search/contact", headers=zi_headers, json={
        "companyName": mc["zi_search"],
        "jobTitle": FINANCE_TITLES_ZI,
        "rpp": 10,
    })
    zi_contacts = zi_resp2.json().get("data", [])
    log.info(f"    ZoomInfo: {len(zi_contacts)} finance contacts")
    time.sleep(0.5)

    # ── Step 4: Apollo Contact Search ──
    log.info("  [4/9] Apollo Contacts...")
    ap_result = apollo._request("POST", "/api/v1/mixed_people/api_search", json_body={
        "q_organization_domains_list": [mc["domain"]],
        "person_titles": FINANCE_TITLES_APOLLO,
        "person_seniorities": ["c_suite", "vp", "director", "owner"],
        "per_page": 10,
    })
    ap_contacts = ap_result.get("people", [])
    log.info(f"    Apollo: {len(ap_contacts)} finance contacts")

    # ── ZoomInfo vs Apollo Comparison ──
    zi_names_set = set()
    zi_contact_map = {}
    for zc in zi_contacts:
        full = f"{zc.get('firstName', '')} {zc.get('lastName', '')}"
        zi_names_set.add(full.lower())
        zi_contact_map[full.lower()] = {
            "name": full,
            "title": zc.get("jobTitle", ""),
            "has_email": zc.get("hasEmail", False),
            "has_phone": zc.get("hasDirectPhone", False),
            "accuracy": zc.get("contactAccuracyScore", ""),
        }

    ap_names_set = set()
    ap_contact_map = {}
    for ac in ap_contacts:
        full = f"{ac.get('first_name', '')} {ac.get('last_name', '')}"
        full_obf = f"{ac.get('first_name', '')} {ac.get('last_name_obfuscated', ac.get('last_name', ''))}"
        ap_names_set.add(full.lower())
        ap_contact_map[full.lower()] = {
            "name": full_obf,
            "title": ac.get("title", ""),
            "apollo_id": ac.get("id", ""),
            "has_email": bool(ac.get("email")),
            "seniority": ac.get("seniority", ""),
        }

    # Contacts only in ZoomInfo (Apollo can't find)
    zi_only = zi_names_set - ap_names_set
    # Contacts only in Apollo
    ap_only = ap_names_set - zi_names_set
    # In both
    both = zi_names_set & ap_names_set

    for name_key in zi_names_set | ap_names_set:
        row = {
            "Company": mc["name"],
            "Contact Name": "",
            "Title": "",
            "In ZoomInfo": "No",
            "In Apollo": "No",
            "Apollo ID": "",
            "ZoomInfo Has Email": "",
            "ZoomInfo Has Phone": "",
            "ZoomInfo Accuracy": "",
            "Apollo Has Email": "",
            "Apollo Seniority": "",
            "Source": "",
        }
        if name_key in zi_contact_map:
            zc = zi_contact_map[name_key]
            row["Contact Name"] = zc["name"]
            row["Title"] = zc["title"]
            row["In ZoomInfo"] = "Yes"
            row["ZoomInfo Has Email"] = "Yes" if zc["has_email"] else "No"
            row["ZoomInfo Has Phone"] = "Yes" if zc["has_phone"] else "No"
            row["ZoomInfo Accuracy"] = str(zc["accuracy"])
        if name_key in ap_contact_map:
            ac = ap_contact_map[name_key]
            if not row["Contact Name"]:
                row["Contact Name"] = ac["name"]
            if not row["Title"]:
                row["Title"] = ac["title"]
            row["In Apollo"] = "Yes"
            row["Apollo ID"] = ac["apollo_id"]
            row["Apollo Has Email"] = "Yes" if ac["has_email"] else "No"
            row["Apollo Seniority"] = ac["seniority"]

        if name_key in both:
            row["Source"] = "Both"
        elif name_key in zi_only:
            row["Source"] = "ZoomInfo Only"
        else:
            row["Source"] = "Apollo Only"

        zi_vs_apollo.append(row)

    log.info(f"    Overlap: {len(both)} both | {len(zi_only)} ZoomInfo-only | {len(ap_only)} Apollo-only")

    # ── Step 4b: Cross-match ZoomInfo-only contacts → Apollo ──
    log.info("  [4b/9] Cross-matching ZoomInfo-only → Apollo...")
    extra_apollo = []
    for name_key in zi_only:
        zc = zi_contact_map[name_key]
        search_q = f"{zc['name']} {mc['name']}"
        result = apollo._request("POST", "/api/v1/mixed_people/api_search", json_body={
            "q_keywords": search_q,
            "per_page": 3,
        })
        matches = result.get("people", [])
        if matches:
            best = matches[0]
            log.info(f"    Cross-matched: {zc['name']} → Apollo ID {best['id']}")
            extra_apollo.append(best)
            # Update comparison row
            for r in zi_vs_apollo:
                if r["Contact Name"] == zc["name"] and r["Company"] == mc["name"]:
                    r["In Apollo"] = "Yes (cross-matched)"
                    r["Apollo ID"] = best["id"]
                    r["Source"] = "Both (cross-matched)"
        else:
            log.info(f"    No Apollo match for: {zc['name']}")
        time.sleep(0.5)

    # Combine all Apollo contacts (deduplicate by ID)
    all_apollo_for_company = []
    seen_ids = set()
    for ac in ap_contacts + extra_apollo:
        aid = ac.get("id", "")
        if aid and aid not in seen_ids:
            seen_ids.add(aid)
            all_apollo_for_company.append(ac)

    if not all_apollo_for_company and not zi_contacts:
        no_prospect_companies.append(mc["name"])
        log.info(f"  ⚠ NO PROSPECTS FOUND for {mc['name']}")
        continue

    # ── Step 5: Apollo Person Enrichment ──
    log.info(f"  [5/9] Enriching {len(all_apollo_for_company)} contacts via Apollo...")
    enriched_contacts = []
    for ac in all_apollo_for_company:
        aid = ac["id"]
        log.info(f"    Enriching {ac.get('first_name', '')} {ac.get('last_name_obfuscated', ac.get('last_name', ''))}...")
        result = apollo.enrich_person(aid)
        person = result.get("person")
        if person:
            e = apollo._extract_person(person)
            e["_source"] = "Apollo"
            e["_apollo_search_title"] = ac.get("title", "")
            enriched_contacts.append(e)
            log.info(f"      → {e['name']} | {e['title']} | {e.get('linkedin_url', 'no LI')}")
        else:
            log.warning(f"      → Enrichment failed for {aid}")
        time.sleep(random.uniform(1, 2))

    # Add ZoomInfo-only contacts that couldn't be cross-matched (limited data)
    for name_key in zi_only:
        zc = zi_contact_map[name_key]
        # Check if we already enriched this person via cross-match
        already = any(e["name"] and e["name"].lower() == name_key for e in enriched_contacts)
        if not already:
            parts = zc["name"].split(None, 1)
            enriched_contacts.append({
                "apollo_id": "",
                "name": zc["name"],
                "first_name": parts[0] if parts else zc["name"],
                "last_name": parts[1] if len(parts) > 1 else "",
                "title": zc["title"],
                "seniority": "",
                "linkedin_url": "",
                "email": "",
                "city": "",
                "state": "",
                "headline": zc["title"],
                "company_name": mc["name"],
                "company_industry": co.get("apollo_industry", ""),
                "company_employees": co.get("apollo_employees", ""),
                "company_revenue": co.get("apollo_revenue", ""),
                "_source": "ZoomInfo Only",
            })

    if not enriched_contacts:
        no_prospect_companies.append(mc["name"])
        log.info(f"  ⚠ NO ENRICHABLE PROSPECTS for {mc['name']}")
        continue

    # ── Step 6: LinkedIn Validation ──
    log.info(f"  [6/9] LinkedIn validation for {len(enriched_contacts)} contacts...")

    # 6a. Company page (one per company)
    company_li_data = {}
    if co.get("company_linkedin"):
        log.info(f"    Company page: {co['company_linkedin']}")
        items = run_actor(COMPANY_SCRAPER, {"companies": [co["company_linkedin"]]})
        if items:
            cp = items[0]
            founded_raw = cp.get("foundedOn", "")
            if isinstance(founded_raw, dict):
                founded_raw = str(founded_raw.get("year", ""))
            company_li_data = {
                "li_followers": cp.get("followerCount", ""),
                "li_employees": cp.get("employeeCount", ""),
                "li_tagline": cp.get("tagline", ""),
                "li_description": (cp.get("description") or "")[:300],
                "li_founded": founded_raw,
                "li_has_logo": "Yes" if cp.get("logo") else "No",
                "li_page_quality": "",
            }
            # Assess quality
            quality_parts = []
            if cp.get("logo"):
                quality_parts.append("logo")
            if cp.get("tagline"):
                quality_parts.append("tagline")
            if cp.get("description"):
                quality_parts.append("description")
            fc = cp.get("followerCount", 0)
            quality_parts.append(f"{fc:,} followers" if fc else "no followers")
            company_li_data["li_page_quality"] = f"{'Good' if len(quality_parts) >= 3 else 'Basic'} — {', '.join(quality_parts)}."
            log.info(f"      {company_li_data['li_page_quality']}")

    # 6b. Per-prospect: profile + posts
    for ec in enriched_contacts:
        li_url = ec.get("linkedin_url", "")
        ec["_li_verified"] = ""
        ec["_li_headline"] = ""
        ec["_li_connections"] = ""
        ec["_li_followers"] = ""
        ec["_li_open_to_work"] = ""
        ec["_li_current_company"] = ""
        ec["_role_verified"] = ""
        ec["_activity_level"] = "Unknown"
        ec["_recent_post_date"] = ""
        ec["_recent_post_text"] = ""
        ec["_posts_count"] = 0
        ec["_reposts_count"] = 0
        ec["_total_feed"] = 0

        if not li_url:
            ec["_activity_level"] = "No LinkedIn URL"
            continue

        slug = li_url.rstrip("/").split("/")[-1]
        log.info(f"    Profile: {ec['name']} ({slug})...")

        # Profile scraper — role verification
        profile_items = run_actor(PROFILE_SCRAPER, {"urls": [li_url]})
        if profile_items:
            p = profile_items[0]
            ec["_li_headline"] = p.get("headline", "")
            ec["_li_connections"] = p.get("connectionsCount", "")
            ec["_li_followers"] = p.get("followerCount", "")
            ec["_li_open_to_work"] = "Yes" if p.get("openToWork") else "No"
            current_positions = p.get("currentPosition") or []
            if current_positions:
                ec["_li_current_company"] = current_positions[0].get("companyName", "")
            # Role verification
            apollo_title = (ec.get("title") or "").lower()
            li_headline = (ec["_li_headline"] or "").lower()
            if apollo_title and li_headline:
                # Check if key words match
                key_words = [w for w in apollo_title.split() if len(w) > 3]
                matches = sum(1 for w in key_words if w in li_headline)
                if matches >= len(key_words) * 0.5:
                    ec["_role_verified"] = "Yes"
                else:
                    ec["_role_verified"] = f"MISMATCH — Apollo: '{ec.get('title')}' vs LinkedIn: '{ec['_li_headline']}'"
            log.info(f"      Headline: {ec['_li_headline'][:60]} | Verified: {ec['_role_verified']}")

        # Posts scraper — activity
        log.info(f"    Activity: {ec['name']}...")
        posts_items = run_actor(POSTS_SCRAPER, {"max_posts": 10, "profiles": [li_url]})
        now = datetime.now()
        activities = []
        slug_lower = slug.lower()

        for item in posts_items:
            author_url = ((item.get("author") or {}).get("profile_url") or "").lower()
            is_authored = slug_lower in author_url
            rb = item.get("repostedBy") or item.get("resharedBy")
            content = (item.get("text") or item.get("content") or "")[:150]
            pd = parse_date((item.get("postedAt") or item.get("posted_at") or {}).get("date", ""))
            if not pd and item.get("date"):
                pd = parse_date(item["date"])

            if rb:
                activities.append({"date": pd, "type": "Repost", "detail": content[:80]})
            elif is_authored:
                activities.append({"date": pd, "type": "Post", "detail": content[:80]})

        activities.sort(key=lambda x: x["date"] or datetime.min, reverse=True)

        ec["_posts_count"] = sum(1 for a in activities if a["type"] == "Post")
        ec["_reposts_count"] = sum(1 for a in activities if a["type"] == "Repost")
        ec["_total_feed"] = len(activities)

        if activities and activities[0]["date"]:
            ld = activities[0]["date"]
            ec["_recent_post_date"] = ld.strftime("%Y-%m-%d")
            ec["_recent_post_text"] = activities[0]["detail"]
            days_ago = (now - ld).days
            if days_ago <= 30:
                ec["_activity_level"] = f"Active — last post {ec['_recent_post_date']}"
            elif days_ago <= 90:
                ec["_activity_level"] = f"Moderate — last post {ec['_recent_post_date']}"
            elif days_ago <= 180:
                ec["_activity_level"] = f"Low — last post {ec['_recent_post_date']}"
            else:
                ec["_activity_level"] = f"Inactive — last post {ec['_recent_post_date']}"
        elif not activities:
            ec["_activity_level"] = "No posts found"

        log.info(f"      {ec['_activity_level']} | Posts: {ec['_posts_count']} Reposts: {ec['_reposts_count']}")
        time.sleep(1)

    # ── Step 7: AI Scoring ──
    log.info(f"  [7/9] AI scoring {len(enriched_contacts)} prospects...")
    # Build scoring input
    scoring_input = []
    for ec in enriched_contacts:
        scoring_input.append({
            "apollo_id": ec.get("apollo_id", ec.get("name", "")),
            "name": ec.get("name", ""),
            "title": ec.get("title", ""),
            "seniority": ec.get("seniority", ""),
            "company_name": ec.get("company_name", mc["name"]),
            "company_industry": ec.get("company_industry", co.get("apollo_industry", "")),
            "company_employees": ec.get("company_employees", co.get("apollo_employees", "")),
            "company_employee_range": ec.get("company_employee_range", ""),
            "company_revenue": ec.get("company_revenue", co.get("apollo_revenue", "")),
            "company_location": co.get("gp_address", ""),
            "city": ec.get("city", ""),
            "state": ec.get("state", ""),
            "linkedin_url": ec.get("linkedin_url", ""),
        })

    # Call scoring via direct OpenAI API (avoid msal import chain)
    SCORE_SYSTEM = """You are an ICP (Ideal Customer Profile) scoring agent for a CPA firm's B2B outreach tool.
You will receive an ICP definition and a list of prospects. Score each prospect 0-100 against the ICP.

Return a JSON object with a "scores" key containing an array where each element has:
- "apollo_id": the prospect's Apollo ID (pass through exactly)
- "score": integer 0-100
- "breakdown": object with dimension scores, e.g. {"title": 25, "industry": 18, "company_size": 20, "location": 10, "seniority": 12, "linkedin": 5, "revenue": 10}
- "reasoning": 1-2 sentence human-readable explanation for why this score

Scoring dimensions and max points:
- title (0-25): CFO/Controller/VP Finance = full marks. Owner/President/CEO = 15-20.
- industry (0-18): Be GENEROUS. "machinery" = manufacturing, "civil engineering" = professional services, etc.
- company_size (0-20): Based on employee count ranges provided.
- revenue (0-12): If missing, score 8/12 (benefit of doubt). Hard exclude >$150M.
- location (0-10): Use company_location (Google Places) if available.
- seniority (0-10): c_suite/vp/director/owner = full marks.
- linkedin (0-5): Has LinkedIn profile?

HARD EXCLUSIONS (score 0): revenue >$150M, government, banking, Fortune 500.
When in doubt, score HIGHER. These are pre-filtered prospects.

Return ONLY valid JSON with {"scores": [...]}."""

    score_user = json.dumps({
        "icp": {
            "target_titles": ICP_CONFIG["target_titles"],
            "target_seniorities": ICP_CONFIG["target_seniorities"],
            "target_industries": ICP_CONFIG["target_industries"],
            "target_locations": ICP_CONFIG["target_locations"],
            "employee_count_ranges": ICP_CONFIG["employee_count_ranges"],
            "revenue_ranges": ICP_CONFIG["scoring_config"]["revenue_ranges"],
            "hard_exclusions": ICP_CONFIG["scoring_config"]["hard_exclusions"],
            "custom_notes": ICP_CONFIG["scoring_config"]["custom_notes"],
        },
        "prospects": scoring_input,
    }, indent=2)

    score_raw = openai_call(SCORE_SYSTEM, score_user, temperature=0.3, max_tokens=2000)
    if score_raw.startswith("```"):
        score_raw = score_raw.split("\n", 1)[1] if "\n" in score_raw else score_raw[3:]
        if score_raw.endswith("```"):
            score_raw = score_raw[:-3].strip()
    score_parsed = json.loads(score_raw)
    if isinstance(score_parsed, dict) and "scores" in score_parsed:
        scores = score_parsed["scores"]
    elif isinstance(score_parsed, list):
        scores = score_parsed
    else:
        scores = []
    log.info(f"    Scored {len(scores)} prospects")

    # Map scores back
    score_map = {}
    for s in scores:
        score_map[s.get("apollo_id", "")] = s

    for ec in enriched_contacts:
        key = ec.get("apollo_id", ec.get("name", ""))
        sc = score_map.get(key, {})
        ec["_icp_score"] = sc.get("score", "")
        ec["_icp_reasoning"] = sc.get("reasoning", "")
        bd = sc.get("breakdown", {})
        ec["_icp_breakdown"] = " | ".join(f"{k}: {v}" for k, v in bd.items()) if bd else ""

    # ── Step 8: Connection Notes + Messages ──
    log.info(f"  [8/9] Generating connection notes + messages...")
    for ec in enriched_contacts:
        first = ec.get("first_name", "")
        if not first:
            continue

        prospect_ctx = json.dumps({
            "prospect": {
                "first_name": first,
                "title": ec.get("title", ""),
                "company_name": ec.get("company_name", mc["name"]),
                "location": co.get("gp_address", ""),
                "industry": ec.get("company_industry", co.get("apollo_industry", "")),
                "headline": ec.get("_li_headline", ec.get("title", "")),
            },
            "company": {
                "industry": co.get("apollo_industry", ""),
                "employees": str(co.get("apollo_employees", "")),
                "revenue": co.get("apollo_revenue", ""),
                "founded": str(co.get("apollo_founded", "")),
                "city": co.get("apollo_city", ""),
            },
        }, indent=2)

        # Connection notes for both senders
        for sender in ["Melinda Johnson", "Adrienne Nordland"]:
            key = f"_note_{sender.split()[0].lower()}"
            try:
                note = openai_call(
                    CONNECTION_NOTE_PROMPT.format(sender_name=sender),
                    prospect_ctx,
                    max_tokens=100,
                )
                note = note.strip().strip('"')
                if len(note) > 200:
                    note = note[:197] + "..."
                ec[key] = note
                log.info(f"    {sender} → {first}: {note[:60]}...")
            except Exception as e:
                ec[key] = ""
                log.warning(f"    Note gen failed for {sender} → {first}: {e}")
            time.sleep(0.5)

        # Messages (3-step sequence)
        for sender in ["Melinda Johnson", "Adrienne Nordland"]:
            key_prefix = f"_msg_{sender.split()[0].lower()}"
            try:
                raw = openai_call(
                    MESSAGE_PROMPT.format(sender_name=sender),
                    prospect_ctx,
                    max_tokens=1000,
                )
                if raw.startswith("```"):
                    raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
                    if raw.endswith("```"):
                        raw = raw[:-3].strip()
                msgs = json.loads(raw)
                for m in msgs:
                    ec[f"{key_prefix}_{m['step']}"] = m["text"]
                log.info(f"    {sender} messages → {first}: 3 generated")
            except Exception as e:
                for i in range(1, 4):
                    ec[f"{key_prefix}_{i}"] = ""
                log.warning(f"    Message gen failed for {sender} → {first}: {e}")
            time.sleep(0.5)

    # ── Build prospect rows ──
    for ec in enriched_contacts:
        row = {
            "ICP Score": ec.get("_icp_score", ""),
            "First Name": ec.get("first_name", ""),
            "Last Name": ec.get("last_name", ""),
            "Title": ec.get("title", ""),
            "Company": ec.get("company_name", mc["name"]),
            "Industry": ec.get("company_industry", co.get("apollo_industry", "")),
            "Employees": ec.get("company_employees", co.get("apollo_employees", "")),
            "Revenue": ec.get("company_revenue", co.get("apollo_revenue", "")),
            "Company City": co.get("apollo_city", ""),
            "Company Domain": ec.get("company_domain", co.get("domain", "")),
            "Email": ec.get("email", ""),
            "Email Status": (ec.get("raw_person") or {}).get("email_status", ""),
            "LinkedIn URL": ec.get("linkedin_url", ""),
            "Seniority": ec.get("seniority", ""),
            "Headline": ec.get("headline", ""),
            "LinkedIn Headline": ec.get("_li_headline", ""),
            "LinkedIn Connections": ec.get("_li_connections", ""),
            "LinkedIn Followers": ec.get("_li_followers", ""),
            "Open to Work": ec.get("_li_open_to_work", ""),
            "LinkedIn Current Company": ec.get("_li_current_company", ""),
            "Role Verified": ec.get("_role_verified", ""),
            "Company LinkedIn URL": co.get("company_linkedin", ""),
            "Company LI Followers": company_li_data.get("li_followers", ""),
            "Company LI Employees": company_li_data.get("li_employees", ""),
            "Company LI Tagline": company_li_data.get("li_tagline", ""),
            "Company LI Description": company_li_data.get("li_description", ""),
            "Company LI Founded": company_li_data.get("li_founded", ""),
            "Company LI Has Logo": company_li_data.get("li_has_logo", ""),
            "Activity Level": ec.get("_activity_level", ""),
            "Recent Post Date": ec.get("_recent_post_date", ""),
            "Recent Post Text": ec.get("_recent_post_text", ""),
            "Posts Count": ec.get("_posts_count", 0),
            "Reposts Count": ec.get("_reposts_count", 0),
            "Total Feed Items": ec.get("_total_feed", 0),
            "ICP Reasoning": ec.get("_icp_reasoning", ""),
            "ICP Score Breakdown": ec.get("_icp_breakdown", ""),
            "Melinda's Connection Note": ec.get("_note_melinda", ""),
            "Adrienne's Connection Note": ec.get("_note_adrienne", ""),
            "Message 1 — Melinda": ec.get("_msg_melinda_1", ""),
            "Message 2 — Melinda": ec.get("_msg_melinda_2", ""),
            "Message 3 — Melinda": ec.get("_msg_melinda_3", ""),
            "Message 1 — Adrienne": ec.get("_msg_adrienne_1", ""),
            "Message 2 — Adrienne": ec.get("_msg_adrienne_2", ""),
            "Message 3 — Adrienne": ec.get("_msg_adrienne_3", ""),
            "Google Places Address": co.get("gp_address", ""),
            "Google Phone": co.get("gp_phone", ""),
            "Google Rating": co.get("gp_rating", ""),
            "Google Reviews": co.get("gp_reviews", ""),
            "Apollo ID": ec.get("apollo_id", ""),
            "Data Source": ec.get("_source", ""),
            "Data Source Pipeline": f"Google Places → Apollo Org → ZoomInfo + Apollo Search → Apollo Enrich → Apify LI → GPT-5.4 Score → Messages",
        }
        all_prospects.append(row)

    company_data_map[mc["name"]] = co
    log.info(f"  ✓ {mc['name']}: {len(enriched_contacts)} prospects added")


# ══════════════════════════════════════════════════════════════
#  OUTPUT: Excel
# ══════════════════════════════════════════════════════════════

OUTFILE = "docs/ICP-Prospects/model_clients_full_pipeline.xlsx"
log.info(f"\n{'=' * 70}")
log.info(f"  WRITING EXCEL: {OUTFILE}")
log.info(f"{'=' * 70}")

wb = openpyxl.Workbook()

# ── Sheet 1: Prospects ──
ws1 = wb.active
ws1.title = "Prospects"

if all_prospects:
    headers = list(all_prospects[0].keys())
    # Header row
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    # Data rows
    for row_idx, prospect in enumerate(all_prospects, 2):
        for col, h in enumerate(headers, 1):
            val = prospect.get(h, "")
            # Convert non-serializable types to string for Excel
            if isinstance(val, (dict, list)):
                val = str(val)
            ws1.cell(row=row_idx, column=col, value=val)

    # Auto-width (approximate)
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = min(30, max(12, len(headers[col - 1]) + 4))

    # Freeze header
    ws1.freeze_panes = "A2"

# ── Sheet 2: ZoomInfo vs Apollo ──
ws2 = wb.create_sheet("ZoomInfo vs Apollo")

if zi_vs_apollo:
    zi_headers = list(zi_vs_apollo[0].keys())
    header_fill2 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col, h in enumerate(zi_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill2
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for row_idx, item in enumerate(zi_vs_apollo, 2):
        for col, h in enumerate(zi_headers, 1):
            val = item.get(h, "")
            if isinstance(val, (dict, list)):
                val = str(val)
            cell = ws2.cell(row=row_idx, column=col, value=val)
            # Highlight ZoomInfo-only rows
            if h == "Source" and "ZoomInfo Only" in str(val):
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif h == "Source" and "Apollo Only" in str(val):
                cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for col in range(1, len(zi_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = min(25, max(12, len(zi_headers[col - 1]) + 4))
    ws2.freeze_panes = "A2"

# ── Sheet 3: No Prospects ──
ws3 = wb.create_sheet("No Prospects Found")
ws3.cell(row=1, column=1, value="Company").font = Font(bold=True)
ws3.cell(row=1, column=2, value="Industry").font = Font(bold=True)
ws3.cell(row=1, column=3, value="Notes").font = Font(bold=True)
ws3.cell(row=1, column=4, value="Reason").font = Font(bold=True)

for i, name in enumerate(no_prospect_companies, 2):
    mc_data = next((m for m in MODEL_CLIENTS if m["name"] == name), {})
    ws3.cell(row=i, column=1, value=name)
    ws3.cell(row=i, column=2, value=mc_data.get("industry", ""))
    ws3.cell(row=i, column=3, value=mc_data.get("notes", ""))
    ws3.cell(row=i, column=4, value="No finance contacts found in ZoomInfo or Apollo. Try manual LinkedIn search or other sources.")

# Also add companies with 0 Apollo+ZoomInfo contacts even if not in no_prospect_companies
for mc in MODEL_CLIENTS:
    if mc["name"] not in no_prospect_companies:
        count = sum(1 for p in all_prospects if p["Company"] == mc["name"])
        if count == 0:
            r = ws3.max_row + 1
            ws3.cell(row=r, column=1, value=mc["name"])
            ws3.cell(row=r, column=2, value=mc["industry"])
            ws3.cell(row=r, column=3, value=mc["notes"])
            ws3.cell(row=r, column=4, value="No enrichable contacts found.")

for col in range(1, 5):
    ws3.column_dimensions[get_column_letter(col)].width = 30

# ── Sheet 4: Summary ──
ws4 = wb.create_sheet("Pipeline Summary")
ws4.cell(row=1, column=1, value="Model Client Pipeline Summary").font = Font(bold=True, size=14)
ws4.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

ws4.cell(row=4, column=1, value="Company").font = Font(bold=True)
ws4.cell(row=4, column=2, value="Industry").font = Font(bold=True)
ws4.cell(row=4, column=3, value="ZoomInfo Contacts").font = Font(bold=True)
ws4.cell(row=4, column=4, value="Apollo Contacts").font = Font(bold=True)
ws4.cell(row=4, column=5, value="ZoomInfo Only").font = Font(bold=True)
ws4.cell(row=4, column=6, value="Apollo Only").font = Font(bold=True)
ws4.cell(row=4, column=7, value="Total Prospects").font = Font(bold=True)
ws4.cell(row=4, column=8, value="Has LinkedIn Data").font = Font(bold=True)

for i, mc in enumerate(MODEL_CLIENTS, 5):
    name = mc["name"]
    zi_count = sum(1 for r in zi_vs_apollo if r["Company"] == name and r["In ZoomInfo"] != "No")
    ap_count = sum(1 for r in zi_vs_apollo if r["Company"] == name and "Yes" in r.get("In Apollo", ""))
    zi_only_count = sum(1 for r in zi_vs_apollo if r["Company"] == name and r["Source"] == "ZoomInfo Only")
    ap_only_count = sum(1 for r in zi_vs_apollo if r["Company"] == name and r["Source"] == "Apollo Only")
    total = sum(1 for p in all_prospects if p["Company"] == name)
    has_li = sum(1 for p in all_prospects if p["Company"] == name and p.get("LinkedIn URL"))

    ws4.cell(row=i, column=1, value=name)
    ws4.cell(row=i, column=2, value=mc["industry"])
    ws4.cell(row=i, column=3, value=zi_count)
    ws4.cell(row=i, column=4, value=ap_count)
    ws4.cell(row=i, column=5, value=zi_only_count)
    ws4.cell(row=i, column=6, value=ap_only_count)
    ws4.cell(row=i, column=7, value=total)
    ws4.cell(row=i, column=8, value=has_li)

for col in range(1, 9):
    ws4.column_dimensions[get_column_letter(col)].width = 22

wb.save(OUTFILE)
log.info(f"\n✓ EXCEL SAVED: {OUTFILE}")
log.info(f"  Prospects: {len(all_prospects)}")
log.info(f"  ZoomInfo vs Apollo comparisons: {len(zi_vs_apollo)}")
log.info(f"  Companies with no prospects: {no_prospect_companies}")

# Also save CSV for compatibility
CSV_OUT = OUTFILE.replace(".xlsx", ".csv")
if all_prospects:
    with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(all_prospects[0].keys()))
        writer.writeheader()
        writer.writerows(all_prospects)
    log.info(f"  CSV also saved: {CSV_OUT}")
